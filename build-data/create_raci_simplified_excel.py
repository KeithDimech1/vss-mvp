#!/usr/bin/env python3
"""
Generate professionally formatted Simplified RACI matrix Excel file from CSV
Includes Comments column
"""

import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Define color scheme (professional RACI colors)
COLORS = {
    'R': PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid'),  # Light green
    'A': PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid'),  # Light orange
    'C': PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid'),  # Light blue
    'I': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),  # Very light green
    'S': PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid'),  # Light peach
    'RA': PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid'), # Light red (dual)
    'header': PatternFill(start_color='002060', end_color='002060', fill_type='solid'),  # Navy blue
    'phase_header': PatternFill(start_color='305496', end_color='305496', fill_type='solid'),  # Medium blue
    'legend_bg': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),  # Light gray
    'comments_bg': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),  # Light yellow
}

# Define fonts
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
PHASE_FONT = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
NORMAL_FONT = Font(name='Calibri', size=10)
BOLD_FONT = Font(name='Calibri', size=10, bold=True)

# Define alignment
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Define borders
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def read_csv_data(csv_path):
    """Read CSV file and return data as list of lists"""
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        return list(reader)

def create_raci_sheet(wb, data):
    """Create the main RACI matrix sheet with Comments column"""
    ws = wb.active
    ws.title = "RACI Matrix (Simplified)"

    # Add header row with Comments column
    headers = ['#', 'Work Package / Task', 'Esri SA', 'Lithodat', 'SGS', 'Comments']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col_idx, value=header)
        cell.fill = COLORS['header']
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # Write legend (rows 1-5)
    legend_data = [
        ['RACI LEGEND:', '', '', '', '', ''],
        ['R = Responsible (does the work)', '', '', '', '', ''],
        ['A = Accountable (final decision authority)', '', '', '', '', ''],
        ['C = Consulted (provides input before decision)', '', '', '', '', ''],
        ['I = Informed (kept updated on progress)', '', '', '', '', ''],
        ['', '', '', '', '', ''],
    ]

    for row_idx, row_data in enumerate(legend_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = COLORS['legend_bg']
            if col_idx == 1 and value:
                cell.font = BOLD_FONT
            else:
                cell.font = NORMAL_FONT
            cell.alignment = LEFT_ALIGN
            cell.border = THIN_BORDER

    # Write data rows (starting at row 8)
    current_row = 8
    for row_data in data[7:]:  # Skip header row in CSV (row 7)
        if not any(row_data):  # Skip completely empty rows
            continue

        # Extend row to include Comments column (empty initially)
        extended_row = list(row_data) + [''] * (6 - len(row_data))
        extended_row = extended_row[:6]  # Ensure exactly 6 columns

        for col_idx, value in enumerate(extended_row, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            cell.alignment = LEFT_ALIGN if col_idx in [1, 2, 6] else CENTER_ALIGN
            cell.border = THIN_BORDER

            # Color code comments column background
            if col_idx == 6:
                cell.fill = COLORS['comments_bg']

        current_row += 1

    # Format phase headers (rows with single-digit work package numbers)
    phase_colors = ['305496', '4472C4', '5B9BD5', '70AD47', 'FFC000', 'C55A11', 'A5A5A5', '7030A0', 'E7E6E6', '44546A', '548235']
    phase_idx = 0

    for row_idx in range(8, current_row):
        wp_number = ws.cell(row=row_idx, column=1).value
        task_name = ws.cell(row=row_idx, column=2).value

        # Check if this is a phase header (single digit, all caps task name)
        if wp_number and str(wp_number).strip().isdigit() and task_name and task_name.isupper():
            phase_color = phase_colors[phase_idx % len(phase_colors)]
            phase_idx += 1

            for col_idx in range(1, 7):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = PatternFill(start_color=phase_color, end_color=phase_color, fill_type='solid')
                cell.font = PHASE_FONT
                cell.alignment = CENTER_ALIGN

    # Color code RACI cells (columns C, D, E - indices 3, 4, 5)
    for row_idx in range(8, current_row):
        for col_idx in [3, 4, 5]:  # Esri SA, Lithodat, SGS columns
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value and str(value).strip() in COLORS:
                cell.fill = COLORS[str(value).strip()]
                cell.font = BOLD_FONT

    # Set column widths
    ws.column_dimensions['A'].width = 8   # #
    ws.column_dimensions['B'].width = 60  # Work Package/Task
    ws.column_dimensions['C'].width = 10  # Esri SA
    ws.column_dimensions['D'].width = 10  # Lithodat
    ws.column_dimensions['E'].width = 10  # SGS
    ws.column_dimensions['F'].width = 80  # Comments

    # Freeze panes (top 7 rows, left 2 columns)
    ws.freeze_panes = 'C8'

    # Add data validation for RACI cells (C, D, E columns)
    raci_validation = DataValidation(type="list", formula1='"R,A,C,I,S"', allow_blank=True)
    raci_validation.error = 'Invalid RACI value'
    raci_validation.errorTitle = 'Invalid Entry'
    raci_validation.prompt = 'Select: R (Responsible), A (Accountable), C (Consulted), I (Informed), S (Support)'
    raci_validation.promptTitle = 'RACI Role'

    ws.add_data_validation(raci_validation)
    raci_validation.add(f'C8:E{current_row}')

    return ws

def create_legend_sheet(wb):
    """Create a detailed legend/guide sheet"""
    ws = wb.create_sheet("Legend & Guide")

    # Title
    ws['A1'] = 'RACI Matrix Legend & User Guide (Simplified Version)'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True)
    ws.merge_cells('A1:E1')

    # RACI Definitions
    ws['A3'] = 'RACI Roles Explained'
    ws['A3'].font = BOLD_FONT
    ws.merge_cells('A3:E3')

    roles = [
        ('R', 'Responsible', 'Does the work to complete the task', 'Light Green'),
        ('A', 'Accountable', 'Final decision authority - ONE per task', 'Light Orange'),
        ('C', 'Consulted', 'Provides input before decision', 'Light Blue'),
        ('I', 'Informed', 'Kept updated on progress', 'Very Light Green'),
        ('S', 'Support', 'Provides supporting assistance', 'Light Peach'),
    ]

    ws['A4'] = 'Role'
    ws['B4'] = 'Name'
    ws['C4'] = 'Definition'
    ws['D4'] = 'Color'
    for cell in ['A4', 'B4', 'C4', 'D4']:
        ws[cell].font = BOLD_FONT
        ws[cell].fill = COLORS['legend_bg']

    for idx, (code, name, definition, color_name) in enumerate(roles, start=5):
        ws[f'A{idx}'] = code
        ws[f'B{idx}'] = name
        ws[f'C{idx}'] = definition
        ws[f'D{idx}'] = color_name
        ws[f'A{idx}'].fill = COLORS[code]
        ws[f'A{idx}'].font = BOLD_FONT
        ws[f'A{idx}'].alignment = CENTER_ALIGN

    # RACI Principles
    ws['A11'] = 'RACI Best Practices'
    ws['A11'].font = BOLD_FONT
    ws.merge_cells('A11:E11')

    principles = [
        '1. Each task MUST have exactly ONE Accountable (A) party',
        '2. Multiple parties can be Responsible (R) for doing the work',
        '3. Consulted (C) parties provide input BEFORE decisions are made',
        '4. Informed (I) parties are updated AFTER decisions are made',
        '5. Avoid "RA" dual accountability - split into separate A and R',
        '6. Keep RACI assignments simple - too many roles creates confusion',
    ]

    for idx, principle in enumerate(principles, start=12):
        ws[f'A{idx}'] = principle
        ws[f'A{idx}'].alignment = LEFT_ALIGN

    # Stakeholder Descriptions
    ws['A20'] = 'Stakeholder Descriptions'
    ws['A20'].font = BOLD_FONT
    ws.merge_cells('A20:E20')

    ws['A22'] = 'Party'
    ws['B22'] = 'Role in Project'
    ws['C22'] = 'Primary Responsibilities'
    for cell in ['A22', 'B22', 'C22']:
        ws[cell].font = BOLD_FONT
        ws[cell].fill = COLORS['legend_bg']

    stakeholders = [
        ('ESRI SA', 'Prime Contractor', 'Commercial, Legal, ArcGIS Infrastructure, Client Relationship, Saudi Compliance'),
        ('Lithodat', 'Technical Subcontractor', 'Geological Data Systems, AI/ML Development, LithoSurfer, Training, Quality'),
        ('SGS', 'Client/End User', 'Requirements, Acceptance Testing, Operational Handover, Final Sign-off'),
    ]

    for idx, (party, role, responsibilities) in enumerate(stakeholders, start=23):
        ws[f'A{idx}'] = party
        ws[f'B{idx}'] = role
        ws[f'C{idx}'] = responsibilities
        ws[f'A{idx}'].font = BOLD_FONT

    # Comments Column Usage
    ws['A28'] = 'Comments Column Usage Guide'
    ws['A28'].font = BOLD_FONT
    ws.merge_cells('A28:E28')

    comments_guide = [
        'The Comments column (yellow background) is for:',
        '• Documenting decisions made during partnership discussions',
        '• Recording clarifications or assumptions',
        '• Noting dependencies or prerequisites',
        '• Tracking status updates or blockers',
        '• Linking to related documents or meetings',
        '• Highlighting risks or concerns',
        '',
        'Example comments:',
        '• "Pending ESRI confirmation on ArcGIS Enterprise license count"',
        '• "Requires SGS data access approval before start"',
        '• "Wayne Noble allocated 0.5 FTE for 3 months"',
        '• "Decision: Hybrid architecture approved 2025-12-30"',
    ]

    for idx, comment in enumerate(comments_guide, start=30):
        ws[f'A{idx}'] = comment
        ws[f'A{idx}'].alignment = LEFT_ALIGN

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 80
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15

    return ws

def create_summary_sheet(wb, data):
    """Create a summary dashboard with statistics"""
    ws = wb.create_sheet("Summary Dashboard")

    # Title
    ws['A1'] = 'GDAC-SA Partnership RACI Matrix (Simplified) - Summary'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True)
    ws.merge_cells('A1:F1')

    # Project Info
    ws['A3'] = 'Project'
    ws['B3'] = 'GDAC-SA Advanced Analytics Platform'
    ws['A4'] = 'Partnership'
    ws['B4'] = 'ESRI Saudi Arabia (Prime) + Lithodat (Subcontractor)'
    ws['A5'] = 'Document'
    ws['B5'] = 'Simplified Responsibility Assignment Matrix (RACI)'
    ws['A6'] = 'Version'
    ws['B6'] = '1.0 - 2025-12-30'

    for cell in ['A3', 'A4', 'A5', 'A6']:
        ws[cell].font = BOLD_FONT

    # Count tasks by work package
    ws['A9'] = 'Work Package Summary'
    ws['A9'].font = BOLD_FONT
    ws.merge_cells('A9:G9')

    ws['A10'] = 'WP#'
    ws['B10'] = 'Phase'
    ws['C10'] = 'Task Count'
    ws['D10'] = 'ESRI Accountable'
    ws['E10'] = 'Lithodat Accountable'
    ws['F10'] = 'SGS Accountable'
    ws['G10'] = 'Joint'

    for cell in ['A10', 'B10', 'C10', 'D10', 'E10', 'F10', 'G10']:
        ws[cell].font = BOLD_FONT
        ws[cell].fill = COLORS['legend_bg']

    # Parse data to count tasks
    work_packages = {}
    current_wp = None
    current_phase = None

    for row in data[7:]:  # Skip header
        if not any(row):
            continue

        wp_num = str(row[0]).strip() if row[0] else ''
        phase = row[1] if len(row) > 1 else ''
        esri = str(row[2]).strip() if len(row) > 2 else ''
        lithodat = str(row[3]).strip() if len(row) > 3 else ''
        sgs = str(row[4]).strip() if len(row) > 4 else ''

        # Check if this is a main phase header
        if wp_num.isdigit() and phase and phase.isupper():
            current_wp = wp_num
            current_phase = phase
            if current_wp not in work_packages:
                work_packages[current_wp] = {
                    'phase': current_phase,
                    'total': 0,
                    'esri_a': 0,
                    'lithodat_a': 0,
                    'sgs_a': 0,
                    'joint': 0
                }
        elif current_wp and wp_num and '.' in wp_num:
            # This is a task
            work_packages[current_wp]['total'] += 1
            if esri == 'A':
                work_packages[current_wp]['esri_a'] += 1
            if lithodat == 'A':
                work_packages[current_wp]['lithodat_a'] += 1
            if sgs == 'A':
                work_packages[current_wp]['sgs_a'] += 1

            # Count joint if multiple parties have R or A
            responsible_count = sum([1 for x in [esri, lithodat, sgs] if x in ['R', 'A']])
            if responsible_count > 1:
                work_packages[current_wp]['joint'] += 1

    # Write work package summary
    row_idx = 11
    total_tasks = 0
    total_esri = 0
    total_lithodat = 0
    total_sgs = 0
    total_joint = 0

    for wp_num in sorted(work_packages.keys(), key=lambda x: int(x) if x.isdigit() else 999):
        wp_data = work_packages[wp_num]
        ws[f'A{row_idx}'] = wp_num
        ws[f'B{row_idx}'] = wp_data['phase']
        ws[f'C{row_idx}'] = wp_data['total']
        ws[f'D{row_idx}'] = wp_data['esri_a']
        ws[f'E{row_idx}'] = wp_data['lithodat_a']
        ws[f'F{row_idx}'] = wp_data['sgs_a']
        ws[f'G{row_idx}'] = wp_data['joint']

        total_tasks += wp_data['total']
        total_esri += wp_data['esri_a']
        total_lithodat += wp_data['lithodat_a']
        total_sgs += wp_data['sgs_a']
        total_joint += wp_data['joint']

        row_idx += 1

    # Totals row
    ws[f'A{row_idx}'] = 'TOTAL'
    ws[f'B{row_idx}'] = ''
    ws[f'C{row_idx}'] = total_tasks
    ws[f'D{row_idx}'] = total_esri
    ws[f'E{row_idx}'] = total_lithodat
    ws[f'F{row_idx}'] = total_sgs
    ws[f'G{row_idx}'] = total_joint

    for cell in [f'A{row_idx}', f'C{row_idx}', f'D{row_idx}', f'E{row_idx}', f'F{row_idx}', f'G{row_idx}']:
        ws[cell].font = BOLD_FONT
        ws[cell].fill = COLORS['legend_bg']

    # Percentage breakdown
    row_idx += 2
    ws[f'A{row_idx}'] = 'Accountable Task Distribution'
    ws[f'A{row_idx}'].font = BOLD_FONT
    ws.merge_cells(f'A{row_idx}:E{row_idx}')

    row_idx += 1
    ws[f'A{row_idx}'] = 'Party'
    ws[f'B{row_idx}'] = 'Accountable Tasks'
    ws[f'C{row_idx}'] = 'Percentage'
    ws[f'D{row_idx}'] = 'Primary Focus'

    for cell in [f'A{row_idx}', f'B{row_idx}', f'C{row_idx}', f'D{row_idx}']:
        ws[cell].font = BOLD_FONT
        ws[cell].fill = COLORS['legend_bg']

    row_idx += 1
    esri_pct = round((total_esri / total_tasks * 100), 1) if total_tasks > 0 else 0
    lithodat_pct = round((total_lithodat / total_tasks * 100), 1) if total_tasks > 0 else 0
    sgs_pct = round((total_sgs / total_tasks * 100), 1) if total_tasks > 0 else 0

    ws[f'A{row_idx}'] = 'ESRI SA'
    ws[f'B{row_idx}'] = total_esri
    ws[f'C{row_idx}'] = f'{esri_pct}%'
    ws[f'D{row_idx}'] = 'Commercial, Legal, ArcGIS, Client Management'

    row_idx += 1
    ws[f'A{row_idx}'] = 'Lithodat'
    ws[f'B{row_idx}'] = total_lithodat
    ws[f'C{row_idx}'] = f'{lithodat_pct}%'
    ws[f'D{row_idx}'] = 'Technical, Geological Data, AI/ML, Training'

    row_idx += 1
    ws[f'A{row_idx}'] = 'SGS'
    ws[f'B{row_idx}'] = total_sgs
    ws[f'C{row_idx}'] = f'{sgs_pct}%'
    ws[f'D{row_idx}'] = 'Acceptance, Testing, Operational Decisions'

    row_idx += 1
    ws[f'A{row_idx}'] = 'Joint Tasks'
    ws[f'B{row_idx}'] = total_joint
    ws[f'C{row_idx}'] = f'{round((total_joint / total_tasks * 100), 1)}%' if total_tasks > 0 else '0%'
    ws[f'D{row_idx}'] = 'Collaborative efforts across parties'

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15

    return ws

def main():
    # Paths
    csv_path = '/Users/keithdimech/Pathway/Dev/Lithodat/Viable Systems Model/VSM-Platform-Project/build-data/GDAC-ESRI-Lithodat-RACI-Matrix-SIMPLIFIED.csv'
    excel_path = '/Users/keithdimech/Pathway/Dev/Lithodat/Viable Systems Model/VSM-Platform-Project/build-data/GDAC-ESRI-Lithodat-RACI-Matrix-SIMPLIFIED-FORMATTED.xlsx'

    print("Reading simplified CSV data...")
    data = read_csv_data(csv_path)

    print("Creating Excel workbook...")
    wb = Workbook()

    print("Creating RACI matrix sheet with Comments column...")
    create_raci_sheet(wb, data)

    print("Creating summary dashboard...")
    create_summary_sheet(wb, data)

    print("Creating legend sheet...")
    create_legend_sheet(wb)

    print(f"Saving Excel file to {excel_path}...")
    wb.save(excel_path)

    print("✅ Simplified Excel file created successfully!")
    print(f"📊 File location: {excel_path}")
    print("\nFeatures included:")
    print("  ✓ Color-coded RACI cells (R=Green, A=Orange, C=Blue, I=Light Green)")
    print("  ✓ Comments column added (light yellow background)")
    print("  ✓ Data validation drop-downs for RACI columns")
    print("  ✓ Frozen panes (top 7 rows, left 2 columns)")
    print("  ✓ Color-coded work package sections (11 phases)")
    print("  ✓ Summary Dashboard tab with task distribution")
    print("  ✓ Legend & Guide tab with stakeholder descriptions")
    print("  ✓ Professional formatting and borders")

if __name__ == '__main__':
    main()
