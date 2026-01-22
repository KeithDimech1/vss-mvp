#!/usr/bin/env python3
"""
Generate professionally formatted RACI matrix Excel file from CSV
"""

import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Define color scheme (professional RACI colors)
COLORS = {
    'R': PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid'),  # Light green
    'A': PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid'),  # Light orange
    'C': PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid'),  # Light blue
    'I': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),  # Very light green
    'S': PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid'),  # Light peach
    'RA': PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid'), # Light red (dual - should be fixed)
    'header': PatternFill(start_color='002060', end_color='002060', fill_type='solid'),  # Navy blue
    'phase_header': PatternFill(start_color='305496', end_color='305496', fill_type='solid'),  # Medium blue
    'legend_bg': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),  # Light gray
}

# Define fonts
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
PHASE_FONT = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
NORMAL_FONT = Font(name='Calibri', size=10)
BOLD_FONT = Font(name='Calibri', size=10, bold=True)

# Define alignment
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Define borders
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def read_csv_data(csv_path):
    """Read CSV file and return data as list of lists"""
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        return list(reader)

def create_raci_sheet(wb, data):
    """Create the main RACI matrix sheet"""
    ws = wb.active
    ws.title = "RACI Matrix"

    # Write data
    for row_idx, row_data in enumerate(data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            cell.alignment = LEFT_ALIGN if col_idx <= 3 or col_idx >= 6 else CENTER_ALIGN
            cell.border = THIN_BORDER

    # Format header row
    for col_idx in range(1, 9):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = COLORS['header']
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN

    # Format legend rows (2-7)
    for row_idx in range(2, 8):
        for col_idx in range(1, 9):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = COLORS['legend_bg']
            if col_idx == 6:  # Legend text
                cell.font = BOLD_FONT

    # Format phase headers (rows with work package numbers like "1", "2", etc.)
    current_phase = None
    phase_colors = ['305496', '4472C4', '5B9BD5', '70AD47', 'FFC000', 'C55A11', 'A5A5A5']
    phase_idx = 0

    for row_idx in range(8, len(data) + 1):
        wp_number = ws.cell(row=row_idx, column=1).value

        # Check if this is a phase header (single digit or empty description)
        if wp_number and str(wp_number).strip() and not str(wp_number).strip().endswith('.'):
            if '.' not in str(wp_number):  # Main phase number
                # This is a main phase header
                current_phase = str(wp_number).strip()
                phase_color = phase_colors[phase_idx % len(phase_colors)]
                phase_idx += 1

                for col_idx in range(1, 9):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = PatternFill(start_color=phase_color, end_color=phase_color, fill_type='solid')
                    cell.font = PHASE_FONT
                    cell.alignment = CENTER_ALIGN

        # Color code RACI cells (columns D and E - indices 4 and 5)
        for col_idx in [4, 5]:  # ESRI SA and Lithodat columns
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value and str(value).strip() in COLORS:
                cell.fill = COLORS[str(value).strip()]
                cell.font = BOLD_FONT

    # Set column widths
    ws.column_dimensions['A'].width = 8   # WP#
    ws.column_dimensions['B'].width = 30  # Phase
    ws.column_dimensions['C'].width = 60  # Task
    ws.column_dimensions['D'].width = 10  # ESRI SA
    ws.column_dimensions['E'].width = 10  # Lithodat
    ws.column_dimensions['F'].width = 70  # Description
    ws.column_dimensions['G'].width = 40  # Deliverable
    ws.column_dimensions['H'].width = 50  # Success Criteria

    # Freeze panes (top 2 rows, left 3 columns)
    ws.freeze_panes = 'D2'

    # Add data validation for RACI cells (D and E columns)
    raci_validation = DataValidation(type="list", formula1='"R,A,C,I,S"', allow_blank=True)
    raci_validation.error = 'Invalid RACI value'
    raci_validation.errorTitle = 'Invalid Entry'
    raci_validation.prompt = 'Select: R (Responsible), A (Accountable), C (Consulted), I (Informed), S (Support)'
    raci_validation.promptTitle = 'RACI Role'

    ws.add_data_validation(raci_validation)
    raci_validation.add(f'D2:E{len(data)}')

    return ws

def create_legend_sheet(wb):
    """Create a detailed legend/guide sheet"""
    ws = wb.create_sheet("Legend & Guide")

    # Title
    ws['A1'] = 'RACI Matrix Legend & User Guide'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True)
    ws.merge_cells('A1:E1')

    # RACI Definitions
    ws['A3'] = 'RACI Roles Explained'
    ws['A3'].font = BOLD_FONT
    ws.merge_cells('A3:E3')

    roles = [
        ('R', 'Responsible', 'Does the work to complete the task', 'Light Green'),
        ('A', 'Accountable', 'Final decision authority - ONE per task', 'Light Orange'),
        ('C', 'Consulted', 'Provides input before decision', 'Light Blue'),
        ('I', 'Informed', 'Kept updated on progress', 'Very Light Green'),
        ('S', 'Support', 'Provides supporting assistance', 'Light Peach'),
    ]

    ws['A4'] = 'Role'
    ws['B4'] = 'Name'
    ws['C4'] = 'Definition'
    ws['D4'] = 'Color'
    for cell in ['A4', 'B4', 'C4', 'D4']:
        ws[cell].font = BOLD_FONT
        ws[cell].fill = COLORS['legend_bg']

    for idx, (code, name, definition, color_name) in enumerate(roles, start=5):
        ws[f'A{idx}'] = code
        ws[f'B{idx}'] = name
        ws[f'C{idx}'] = definition
        ws[f'D{idx}'] = color_name
        ws[f'A{idx}'].fill = COLORS[code]
        ws[f'A{idx}'].font = BOLD_FONT
        ws[f'A{idx}'].alignment = CENTER_ALIGN

    # RACI Principles
    ws['A11'] = 'RACI Best Practices'
    ws['A11'].font = BOLD_FONT
    ws.merge_cells('A11:E11')

    principles = [
        '1. Each task MUST have exactly ONE Accountable (A) party',
        '2. Multiple parties can be Responsible (R) for doing the work',
        '3. Consulted (C) parties provide input BEFORE decisions are made',
        '4. Informed (I) parties are updated AFTER decisions are made',
        '5. Avoid "RA" dual accountability - split into separate A and R',
        '6. Keep RACI assignments simple - too many roles creates confusion',
    ]

    for idx, principle in enumerate(principles, start=12):
        ws[f'A{idx}'] = principle
        ws[f'A{idx}'].alignment = LEFT_ALIGN

    # Partnership Allocation Summary
    ws['A20'] = 'Partnership Work Allocation Summary'
    ws['A20'].font = BOLD_FONT
    ws.merge_cells('A20:E20')

    ws['A22'] = 'Party'
    ws['B22'] = 'Accountable (A) Tasks'
    ws['C22'] = 'Percentage'
    ws['D22'] = 'Focus Areas'
    for cell in ['A22', 'B22', 'C22', 'D22']:
        ws[cell].font = BOLD_FONT
        ws[cell].fill = COLORS['legend_bg']

    allocation = [
        ('ESRI SA', '~150 tasks', '43%', 'Commercial, Legal, Insurance, ArcGIS Infrastructure, L1 Support'),
        ('Lithodat', '~160 tasks', '45%', 'Technical, Geological Data, AI/ML, Training, L3 Support'),
        ('Joint (R/R)', '~40 tasks', '12%', 'Planning, Testing, Governance, Risk Management'),
    ]

    for idx, (party, tasks, pct, focus) in enumerate(allocation, start=23):
        ws[f'A{idx}'] = party
        ws[f'B{idx}'] = tasks
        ws[f'C{idx}'] = pct
        ws[f'D{idx}'] = focus
        ws[f'A{idx}'].font = BOLD_FONT

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 15

    return ws

def create_summary_sheet(wb, data):
    """Create a summary dashboard with statistics"""
    ws = wb.create_sheet("Summary Dashboard")

    # Title
    ws['A1'] = 'GDAC-SA Partnership RACI Matrix - Summary Dashboard'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True)
    ws.merge_cells('A1:F1')

    # Project Info
    ws['A3'] = 'Project'
    ws['B3'] = 'GDAC-SA Advanced Analytics Platform'
    ws['A4'] = 'Partnership'
    ws['B4'] = 'ESRI Saudi Arabia (Prime) + Lithodat (Subcontractor)'
    ws['A5'] = 'Document'
    ws['B5'] = 'Responsibility Assignment Matrix (RACI)'
    ws['A6'] = 'Version'
    ws['B6'] = '1.0 - 2025-12-30'

    for cell in ['A3', 'A4', 'A5', 'A6']:
        ws[cell].font = BOLD_FONT

    # Count tasks by work package
    ws['A9'] = 'Work Package Summary'
    ws['A9'].font = BOLD_FONT
    ws.merge_cells('A9:F9')

    ws['A10'] = 'WP#'
    ws['B10'] = 'Phase'
    ws['C10'] = 'Task Count'
    ws['D10'] = 'ESRI Accountable'
    ws['E10'] = 'Lithodat Accountable'
    ws['F10'] = 'Joint'

    for cell in ['A10', 'B10', 'C10', 'D10', 'E10', 'F10']:
        ws[cell].font = BOLD_FONT
        ws[cell].fill = COLORS['legend_bg']

    # Parse data to count tasks
    work_packages = {}
    current_wp = None
    current_phase = None

    for row in data[8:]:  # Skip header and legend
        if not row[0]:  # Skip empty rows
            continue

        wp_num = str(row[0]).strip()
        phase = row[1]
        esri = str(row[3]).strip() if len(row) > 3 else ''
        lithodat = str(row[4]).strip() if len(row) > 4 else ''

        # Check if this is a main phase header
        if wp_num and '.' not in wp_num and phase:
            current_wp = wp_num
            current_phase = phase
            if current_wp not in work_packages:
                work_packages[current_wp] = {
                    'phase': current_phase,
                    'total': 0,
                    'esri_a': 0,
                    'lithodat_a': 0,
                    'joint': 0
                }
        elif current_wp and wp_num and '.' in wp_num:
            # This is a task
            work_packages[current_wp]['total'] += 1
            if esri == 'A':
                work_packages[current_wp]['esri_a'] += 1
            if lithodat == 'A':
                work_packages[current_wp]['lithodat_a'] += 1
            if esri in ['R', 'A'] and lithodat in ['R', 'A']:
                work_packages[current_wp]['joint'] += 1

    # Write work package summary
    row_idx = 11
    total_tasks = 0
    total_esri = 0
    total_lithodat = 0
    total_joint = 0

    for wp_num in sorted(work_packages.keys(), key=lambda x: int(x) if x.isdigit() else 999):
        wp_data = work_packages[wp_num]
        ws[f'A{row_idx}'] = wp_num
        ws[f'B{row_idx}'] = wp_data['phase']
        ws[f'C{row_idx}'] = wp_data['total']
        ws[f'D{row_idx}'] = wp_data['esri_a']
        ws[f'E{row_idx}'] = wp_data['lithodat_a']
        ws[f'F{row_idx}'] = wp_data['joint']

        total_tasks += wp_data['total']
        total_esri += wp_data['esri_a']
        total_lithodat += wp_data['lithodat_a']
        total_joint += wp_data['joint']

        row_idx += 1

    # Totals row
    ws[f'A{row_idx}'] = 'TOTAL'
    ws[f'B{row_idx}'] = ''
    ws[f'C{row_idx}'] = total_tasks
    ws[f'D{row_idx}'] = total_esri
    ws[f'E{row_idx}'] = total_lithodat
    ws[f'F{row_idx}'] = total_joint

    for cell in [f'A{row_idx}', f'C{row_idx}', f'D{row_idx}', f'E{row_idx}', f'F{row_idx}']:
        ws[cell].font = BOLD_FONT
        ws[cell].fill = COLORS['legend_bg']

    # Percentage breakdown
    row_idx += 2
    ws[f'A{row_idx}'] = 'Effort Allocation Breakdown'
    ws[f'A{row_idx}'].font = BOLD_FONT
    ws.merge_cells(f'A{row_idx}:F{row_idx}')

    row_idx += 1
    ws[f'A{row_idx}'] = 'Party'
    ws[f'B{row_idx}'] = 'Accountable Tasks'
    ws[f'C{row_idx}'] = 'Percentage'
    ws[f'D{row_idx}'] = 'Recommended Revenue Split'

    for cell in [f'A{row_idx}', f'B{row_idx}', f'C{row_idx}', f'D{row_idx}']:
        ws[cell].font = BOLD_FONT
        ws[cell].fill = COLORS['legend_bg']

    row_idx += 1
    esri_pct = round((total_esri / total_tasks * 100), 1) if total_tasks > 0 else 0
    lithodat_pct = round((total_lithodat / total_tasks * 100), 1) if total_tasks > 0 else 0
    joint_pct = round((total_joint / total_tasks * 100), 1) if total_tasks > 0 else 0

    ws[f'A{row_idx}'] = 'ESRI SA'
    ws[f'B{row_idx}'] = total_esri
    ws[f'C{row_idx}'] = f'{esri_pct}%'
    ws[f'D{row_idx}'] = '45% (manages commercial/legal overhead)'

    row_idx += 1
    ws[f'A{row_idx}'] = 'Lithodat'
    ws[f'B{row_idx}'] = total_lithodat
    ws[f'C{row_idx}'] = f'{lithodat_pct}%'
    ws[f'D{row_idx}'] = '45% (delivers technical work)'

    row_idx += 1
    ws[f'A{row_idx}'] = 'Joint (R/R)'
    ws[f'B{row_idx}'] = total_joint
    ws[f'C{row_idx}'] = f'{joint_pct}%'
    ws[f'D{row_idx}'] = '10% (overhead/contingency)'

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15

    return ws

def main():
    # Paths
    csv_path = '/Users/keithdimech/Pathway/Dev/Lithodat/Viable Systems Model/VSM-Platform-Project/build-data/GDAC-ESRI-Lithodat-RACI-Matrix-COMPLETE.csv'
    excel_path = '/Users/keithdimech/Pathway/Dev/Lithodat/Viable Systems Model/VSM-Platform-Project/build-data/GDAC-ESRI-Lithodat-RACI-Matrix-FORMATTED.xlsx'

    print("Reading CSV data...")
    data = read_csv_data(csv_path)

    print("Creating Excel workbook...")
    wb = Workbook()

    print("Creating RACI matrix sheet...")
    create_raci_sheet(wb, data)

    print("Creating summary dashboard...")
    create_summary_sheet(wb, data)

    print("Creating legend sheet...")
    create_legend_sheet(wb)

    print(f"Saving Excel file to {excel_path}...")
    wb.save(excel_path)

    print("✅ Excel file created successfully!")
    print(f"📊 File location: {excel_path}")
    print("\nFeatures included:")
    print("  ✓ Color-coded RACI cells (R=Green, A=Orange, C=Blue, I=Light Green)")
    print("  ✓ Data validation drop-downs for RACI columns")
    print("  ✓ Frozen panes (top 2 rows, left 3 columns)")
    print("  ✓ Color-coded work package sections")
    print("  ✓ Summary Dashboard tab with effort allocation")
    print("  ✓ Legend & Guide tab with RACI definitions")
    print("  ✓ Professional formatting and borders")

if __name__ == '__main__':
    main()
